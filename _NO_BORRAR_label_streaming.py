"""
label_streaming.py — etiqueta una bitácora Integranot aplicando las reglas
de Estatus y Tramite del proyecto Notaría 31.

Uso:
    python3 label_streaming.py <ruta_bitacora.xlsx>

Comportamiento:
  - Crea respaldo .bak antes de tocar el original.
  - Aplica todas las reglas (traslativas / no traslativas, COO->LBR,
    excepción APLICACIÓN DE BIENES, ENT cierre, cancelaciones, etc.).
  - Solo etiqueta filas con Estatus VACÍO (no sobreescribe).
  - Lee con openpyxl read_only (streaming) y escribe con openpyxl normal
    preservando estilos en columnas no-tocadas.
"""
import sys, shutil, time
from datetime import datetime, date, timedelta
from pathlib import Path
import openpyxl

# ---------- Helpers ----------
def _parse_str_date(s):
    if not isinstance(s, str): return None
    s = s.strip()
    if not s or s == '-' or s == '*': return None
    for fmt in ('%d/%m/%Y','%Y-%m-%d','%d-%m-%Y'):
        try: return datetime.strptime(s, fmt)
        except: pass
    return None
def is_dt(v):
    if isinstance(v,(datetime,date)): return True
    return _parse_str_date(v) is not None
def to_dt(v):
    if isinstance(v, datetime): return v
    if isinstance(v, date): return datetime.combine(v, datetime.min.time())
    return _parse_str_date(v)
def is_dash(v): return isinstance(v, str) and v.strip() == "-"
def is_star(v): return isinstance(v, str) and v.strip() == "*"
def is_empty(v):
    if v is None: return True
    if isinstance(v, str) and v.strip() == "": return True
    return False
def has_date(v): return is_dt(v)
def no_date(v): return not is_dt(v)
def pending_no_dash(v): return is_empty(v) or is_star(v)

def _cancel_rpp(r):
    return is_dash(r.get("RPP. Regreso de RPP Testimonio")) and is_dash(r.get("RPP. Captura Dato Inscripción RPP"))
def _cancel_rpc(r):
    return is_dash(r.get("RPC. Entrega a Unidad RPC"))
def _rpp_apl(r):
    s = r.get("RPP. Solicitud")
    if is_dash(s) or not is_dt(s): return False
    return not _cancel_rpp(r)
def _rpc_apl(r):
    s = r.get("RPC. Solicitud")
    if is_dash(s) or not is_dt(s): return False
    return not _cancel_rpc(r)
def _aplic_bienes(r):
    op = r.get("Operación")
    if not isinstance(op, str): return False
    u = op.upper()
    return "APLICACIÓN DE BIENES" in u or "APLICACION DE BIENES" in u

# ---------- Reglas ----------
def label_no_tras(r):
    cierre_exp = r.get("CIERRE. Expedición de Testimonio")
    ent_dig = r.get("ENT. Testimonio Digitalizado")
    ent_fin = r.get("ENT. Entrega Test a Financiero")
    ent_rec = r.get("ENT. Test. a Recepción")
    rpp_sol = r.get("RPP. Solicitud")
    rpc_sol = r.get("RPC. Solicitud")
    rpp_regreso = r.get("RPP. Regreso de RPP Testimonio")
    rpp_captura = r.get("RPP. Captura Dato Inscripción RPP")
    rpp_entrada = r.get("RPP. Entrada a RPP Testimonio")
    rpp_dig_entrada = r.get("RPP. Digitalización Entrada RPP")
    siger_dig = r.get("SIGER. Dig. Const. Inscripción RPC")
    rpp_no = is_dash(rpp_sol) or (is_dt(rpp_sol) and _cancel_rpp(r))
    rpc_no = is_dash(rpc_sol) or (is_dt(rpc_sol) and _cancel_rpc(r))
    cierre_or_ent = any(has_date(x) for x in [cierre_exp, ent_dig, ent_fin, ent_rec])
    if rpp_no and rpc_no and cierre_or_ent: return ("CONCLUIDO", None)
    if rpp_no and rpc_no and not cierre_or_ent: return ("FALTA CIERRE", None)
    if _rpp_apl(r) and (has_date(rpp_regreso) or has_date(rpp_captura)): return ("CONCLUIDO", None)
    if _rpp_apl(r) and no_date(rpp_regreso) and no_date(rpp_captura):
        if has_date(rpp_entrada) or has_date(rpp_dig_entrada): return ("EN RPP", None)
    if _rpp_apl(r) and no_date(rpp_regreso) and no_date(rpp_captura):
        if no_date(rpp_entrada) and no_date(rpp_dig_entrada): return ("FALTA RPP", None)
    if _rpc_apl(r) and has_date(siger_dig): return ("CONCLUIDO", None)
    if _rpc_apl(r) and is_dash(siger_dig) and cierre_or_ent: return ("CONCLUIDO", None)
    if _rpc_apl(r) and pending_no_dash(siger_dig): return ("FALTA SIGER", None)
    if has_date(rpp_regreso) or has_date(rpp_captura) or has_date(siger_dig): return ("CONCLUIDO", None)
    if (is_dash(rpp_regreso) or is_dash(siger_dig)) and cierre_or_ent: return ("CONCLUIDO", None)
    # Fallback: RPP/RPC en estado indeterminado (*) — determinar por cierre/entrega
    if cierre_or_ent: return ("CONCLUIDO", None)
    return ("FALTA CIERRE", None)

def label_tras(r):
    # Caso especial: RPP no aplica para esta traslativa (Solicitud = "-")
    rpp_sol_tr = r.get("RPP. Solicitud")
    if is_dash(rpp_sol_tr):
        cat_notif = r.get("CAT. Dig Notificacion Catastral")
        cierre_or_ent_tr = any(has_date(r.get(c)) for c in [
            'CIERRE. Expedición de Testimonio', 'ENT. Testimonio Digitalizado',
            'ENT. Entrega Test a Financiero', 'ENT. Test. a Recepción'])
        notif_ok = has_date(cat_notif) or is_dash(cat_notif)
        if notif_ok and cierre_or_ent_tr:
            return ("CONCLUIDO", None)
        if not notif_ok and cierre_or_ent_tr:
            return ("FALTA FINALIZAR", "NOTIF CAT")
        # Sin cierre/ENT: determinar por estado de TD y CAT
        td_dig_pago_tr = r.get("TD. Digitalización Pago TD")
        td_pago_tr    = r.get("TD. Pago")
        td_pagado_tr  = has_date(td_dig_pago_tr) or has_date(td_pago_tr)
        if not td_pagado_tr:
            return ("FALTA TD", None)
        if not notif_ok:
            return ("FALTA FINALIZAR", "NOTIF CAT")
        return ("CONCLUIDO", None)
    cat_dig_notif = r.get("CAT. Dig Notificacion Catastral")
    rpp_regreso = r.get("RPP. Regreso de RPP Testimonio")
    rpp_captura = r.get("RPP. Captura Dato Inscripción RPP")
    rpp_entrada = r.get("RPP. Entrada a RPP Testimonio")
    rpp_dig_entrada = r.get("RPP. Digitalización Entrada RPP")
    td_dig_pago = r.get("TD. Digitalización Pago TD")
    td_pago = r.get("TD. Pago")
    cat_dig_entrada = r.get("CAT. Digitalización Entrada a Cat.")
    rpp_term = has_date(rpp_regreso) or has_date(rpp_captura)
    rpp_entry = has_date(rpp_entrada) or has_date(rpp_dig_entrada)
    td_pagado = has_date(td_dig_pago) or has_date(td_pago)
    if is_dash(cat_dig_notif) and rpp_term: return ("CONCLUIDO", None)
    if has_date(cat_dig_notif) and rpp_term: return ("CONCLUIDO", None)
    if no_date(cat_dig_notif) and rpp_term: return ("FALTA FINALIZAR", "NOTIF CAT")
    if has_date(cat_dig_notif) and rpp_entry and no_date(rpp_regreso) and no_date(rpp_captura): return ("EN RPP", None)
    if has_date(cat_dig_notif) and not rpp_entry: return ("FALTA RPP", None)
    if no_date(cat_dig_notif) and rpp_entry: return ("EN RPP", "NOTIF CAT")
    if td_pagado and no_date(cat_dig_notif) and not rpp_entry: return ("FALTA RPP", "NOTIF CAT")
    if (not td_pagado and no_date(cat_dig_notif) and not rpp_entry):
        if _aplic_bienes(r) and has_date(cat_dig_notif): pass
        else: return ("FALTA TD", None)
    return (None, None)

# Sub-procesos que disparan FALTA FINALIZAR cuando una escritura ya
# fue marcada CONCLUIDO pero todavía tiene pasos sin cumplir.
# Vacío sin '-' = pendiente. '-' = no aplica.
CONCILIAR_COLS = [
    'ENT. Test. a Recepción',
    'ENT. Entrega Test a Financiero',
]

def _step_pending(v):
    """True si el sub-paso está pendiente (vacío sin '-')."""
    return is_empty(v) and not is_dash(v)

def _notif_cat_aplica(r):
    """NOTIF CAT: traslativa con CAT. Dig Notificacion Catastral no cumplido."""
    lleva = r.get('Lleva TD')
    is_tr = isinstance(lleva, str) and lleva.strip().upper() == 'SI'
    if not is_tr: return False
    cat_notif = r.get('CAT. Dig Notificacion Catastral')
    return not has_date(cat_notif) and not is_dash(cat_notif)

def _digitalizar_aplica(r):
    """
    DIGITALIZAR aplica cuando:
      - ENT. Testimonio Digitalizado está vacío (no cumplido)
      - CIERRE. Expedición de Testimonio está cumplido (tiene fecha)
      - Si NO traslativa: RPP.Solicitud y RPC.Solicitud no aplican ('-' o vacíos)
      - Si traslativa: RPP. Regreso de RPP Testimonio está cumplido
    """
    ent_dig    = r.get('ENT. Testimonio Digitalizado')
    cierre_exp = r.get('CIERRE. Expedición de Testimonio')
    if not _step_pending(ent_dig): return False
    if not has_date(cierre_exp):   return False
    lleva = r.get('Lleva TD')
    is_tr = isinstance(lleva, str) and lleva.strip().upper() == 'SI'
    if is_tr:
        return has_date(r.get('RPP. Regreso de RPP Testimonio'))
    else:
        rpp_no = is_dash(r.get('RPP. Solicitud')) or is_empty(r.get('RPP. Solicitud'))
        rpc_no = is_dash(r.get('RPC. Solicitud')) or is_empty(r.get('RPC. Solicitud'))
        return rpp_no and rpc_no

def _conciliar_aplica(r):
    """CONCILIAR: cualquiera de las columnas de entrega pendiente."""
    return any(_step_pending(r.get(c)) for c in CONCILIAR_COLS)

def _enviar_aplica(r):
    """ENVIAR: solicitud activa (tiene fecha) pero acuse de recibo pendiente.
    Aplica si cualquiera de las 3 circunstancias se cumple; permanece activa
    hasta que TODAS las que aplican estén resueltas.
    """
    condiciones = [
        (r.get('Solicitud de Entrega de Testimonio al Banco'),
         r.get('BANCOS. Dig. Acuse de Recibo Banco')),
        (r.get('INFONAVIT. Solicitud de Entrega de Testimonio al Infonavit'),
         r.get('INFONAVIT. Dig. Acuse de Recibo INFONAVIT')),
        (r.get('Solicitud Entrega de Testimonio al Fovissste'),
         r.get('FOVISSSTE. Dig. Acuse de Recibo FOVISSSTE')),
    ]
    return any(has_date(sol) and _step_pending(acuse) for sol, acuse in condiciones)

def infer_falta_finalizar_tramite(r):
    """Evalúa todas las subetiquetas aplicables y las combina con '-'.
    Orden: NOTIF CAT - DIGITALIZAR - CONCILIAR - ENVIAR.
    CONCILIAR se suprime si DIGITALIZAR está activo (depende de ella).
    Devuelve None si ninguna aplica.
    """
    partes = []
    if _notif_cat_aplica(r):   partes.append('NOTIF CAT')
    dig = _digitalizar_aplica(r)
    if dig:                    partes.append('DIGITALIZAR')
    if _conciliar_aplica(r) and not dig: partes.append('CONCILIAR')
    if _enviar_aplica(r):      partes.append('ENVIAR')
    return '-'.join(partes) if partes else None

def infer_falta_cierre_tramite(r):
    """PASAR / ARMADO / FOLIO para escrituras con FALTA CIERRE (mutuamente excluyentes).
    PASAR  -> MDC. Entrega Exp Unidad Gestion no cumplido.
    ARMADO -> MDC cumplido y firmas cumplidas (expediente listo, esperando folio).
    FOLIO  -> MDC cumplido pero firmas pendientes.
    """
    mdc    = r.get('MDC. Entrega Exp Unidad Gestión')
    firmas = r.get('ESC. Entrega Firmas Completas')
    if not has_date(mdc):
        return 'PASAR'
    if has_date(firmas):
        return 'ARMADO'
    return 'FOLIO'

def infer_falta_td_tramite(r):
    """PASAR / CAPTURA / RECURSOS para escrituras con FALTA TD.
    PASAR    -> MDC. Entrega Exp Unidad Gestion no cumplido
    CAPTURA  -> MDC cumplido, TD. Elab. Formato y/o Carga portal no cumplido
    RECURSOS -> MDC cumplido y TD. Elab. cumplido
    * se trata como no cumplido (has_date devuelve False para *).
    """
    mdc = r.get('MDC. Entrega Exp Unidad Gestión')
    td_elab = r.get('TD. Elab. Formato y/o Carga portal')
    if not has_date(mdc):
        return 'PASAR'
    if not has_date(td_elab):
        return 'CAPTURA'
    return 'RECURSOS'

def infer_rpp_siger_tramite(r):
    """FOLIO / CIERRE / NOTIF CAT / PAGO / ENTRADA para escrituras con FALTA RPP / FALTA SIGER.
    Prioridad traslativas : FOLIO > CIERRE > NOTIF CAT > PAGO > ENTRADA.
    Prioridad no-traslativas: FOLIO > CIERRE > PAGO > ENTRADA.
      FOLIO    -> ESC. Entrega Firmas Completas sin fecha (prioridad máxima).
      CIERRE   -> Firmas cumplidas pero CIERRE. Expedición de Testimonio sin fecha.
      NOTIF CAT-> (solo traslativas) Firmas y cierre cumplidos, pero CAT.Dig.Notif pendiente.
      PAGO     -> Firmas y cierre cumplidos (y cat ok) pero ningún campo de pago tiene fecha.
      ENTRADA  -> Firmas ✓ + Cierre ✓ + CAT ok + Pago hecho (listo para ingresar al RPP).
    """
    firmas = r.get('ESC. Entrega Firmas Completas')
    cierre = r.get('CIERRE. Expedición de Testimonio')
    pago_cols = [
        'RPP. Pago RPP',
        'RPP. Digitalización Pago RPP',
        'SIGER. Pago RPC',
        'SIGER. Digitalización de Pago RPC',
    ]
    pago_hecho = any(has_date(r.get(c)) for c in pago_cols)

    # NOTIF CAT aplica si: traslativa + CAT.Solicitud activo + CAT.Dig.Notif pendiente
    lleva_td  = str(r.get('Lleva TD', '') or '').strip()
    cat_sol   = r.get('CAT. Solicitud')
    cat_dig   = r.get('CAT. Dig Notificacion Catastral')
    cat_notif_pend = (
        lleva_td == 'SI' and
        has_date(cat_sol) and
        no_date(cat_dig) and not is_dash(cat_dig)
    )

    if not has_date(firmas):
        return 'FOLIO'
    if not has_date(cierre):
        return 'CIERRE'
    if cat_notif_pend:
        return 'NOTIF CAT'
    if not pago_hecho:
        return 'PAGO'
    return 'ENTRADA'

def label_row(r):
    """Devuelve (Estatus, Tramite) para una fila dict-like."""
    e = r.get("Estatus")
    if not is_empty(e) and not is_star(e):
        # Re-evaluar CONCLUIDO y FALTA FINALIZAR para reflejar cambios en
        # PASAR_COLS / CONCILIAR_COLS sin necesidad de borrar etiquetas manualmente.
        if e in ("CONCLUIDO", "FALTA FINALIZAR"):
            sub_tr = infer_falta_finalizar_tramite(r)
            if sub_tr:
                return "FALTA FINALIZAR", sub_tr
            # Si ya no hay sub-paso pendiente, promover a CONCLUIDO
            if e == "FALTA FINALIZAR" and not sub_tr:
                return "CONCLUIDO", ''
        if e in ("FALTA RPP", "FALTA SIGER"):
            return e, infer_rpp_siger_tramite(r)
        if e == "FALTA TD":
            return e, infer_falta_td_tramite(r)
        if e == "FALTA CIERRE":
            return e, infer_falta_cierre_tramite(r)
        # EN RPP: caer al re-etiquetado completo (puede haber concluido)
        if e not in ('EN RPP',):
            return e, r.get("Tramite") or ''
    lleva = r.get("Lleva TD")
    is_tr = isinstance(lleva, str) and lleva.strip().upper() == 'SI'
    est, tr = (label_tras(r) if is_tr else label_no_tras(r))
    if est is None: return '', ''
    # Si quedó CONCLUIDO, revisar si en realidad le falta un sub-paso
    if est == "CONCLUIDO":
        sub_tr = infer_falta_finalizar_tramite(r)
        if sub_tr:
            return "FALTA FINALIZAR", sub_tr
    if est == "CONCLUIDO" and tr: est = "FALTA FINALIZAR"
    # Asignar PAGO o ENTRADA para FALTA RPP y FALTA SIGER
    if est in ("FALTA RPP", "FALTA SIGER"):
        return est, infer_rpp_siger_tramite(r)
    # Asignar PASAR / CAPTURA / RECURSOS para FALTA TD
    if est == "FALTA TD":
        return est, infer_falta_td_tramite(r)
    # Asignar FOLIO para FALTA CIERRE
    if est == "FALTA CIERRE":
        return est, infer_falta_cierre_tramite(r)
    return est, (tr or '')

# ---------- Main ----------
def main():
    if len(sys.argv) < 2:
        print("Uso: python3 label_streaming.py <ruta_bitacora.xlsx>")
        sys.exit(1)
    src = Path(sys.argv[1])
    if not src.exists():
        print(f"ERROR: no existe {src}")
        sys.exit(2)

    # Respaldo .bak
    bak = src.with_suffix(src.suffix + ".bak")
    if not bak.exists():
        shutil.copy(src, bak)
        print(f"Respaldo creado: {bak.name}")
    else:
        print(f"Respaldo ya existía: {bak.name}")

    t0 = time.time()
    print(f"Leyendo {src.name}...")
    wb = openpyxl.load_workbook(src)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    H = {h: i for i, h in enumerate(headers) if h is not None}
    if 'Estatus' not in H or 'Tramite' not in H:
        print(f"ERROR: faltan columnas Estatus/Tramite. Encontradas: {list(H.keys())[:30]}")
        sys.exit(3)
    col_es = H['Estatus'] + 1
    col_tr = H['Tramite'] + 1
    col_ab = H.get('Abogado')

    total = 0
    etiquetadas = 0
    coo_norm = 0
    j31_seen = 0
    for row_idx in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=row_idx, column=i+1).value for i in range(len(headers))]
        if all(v is None for v in row_vals): continue

        # Ignorar filas de totales / basura exportadas por Integranot
        # (celdas con valores numéricos en campos que deberían ser texto/fecha)
        _es_raw = row_vals[col_es - 1] if col_es - 1 < len(row_vals) else None
        _es_test = str(_es_raw).strip() if _es_raw is not None else ''
        VALID_ES_SET = {'CONCLUIDO','FALTA FINALIZAR','FALTA RPP','FALTA SIGER','EN RPP','FALTA TD','FALTA CIERRE',''}
        if _es_test and _es_test not in VALID_ES_SET: continue

        total += 1
        r = dict(zip(headers, row_vals))

        # Normalizar Abogado COO->LBR
        if col_ab is not None:
            ab = r.get('Abogado')
            if isinstance(ab, str):
                ab_u = ab.strip().upper()
                if ab_u == 'COO':
                    ws.cell(row=row_idx, column=col_ab+1).value = 'LBR'
                    r['Abogado'] = 'LBR'
                    coo_norm += 1
                elif ab_u == 'J31':
                    j31_seen += 1

        # Etiquetar si Estatus está vacío. Adicionalmente re-evaluar filas
        # CONCLUIDO y FALTA FINALIZAR para que cambios en las listas
        # PASAR_COLS / CONCILIAR_COLS se reflejen en corridas posteriores.
        es_actual = r.get('Estatus')
        es_actual_str = es_actual.strip() if isinstance(es_actual, str) else ''
        RE_EVAL = ('CONCLUIDO', 'FALTA FINALIZAR', 'FALTA RPP', 'FALTA SIGER', 'FALTA TD', 'FALTA CIERRE', 'EN RPP')
        if (not is_empty(es_actual) and not is_star(es_actual)
                and es_actual_str not in RE_EVAL):
            continue

        est, tr = label_row(r)
        if est:
            ws.cell(row=row_idx, column=col_es).value = est
            ws.cell(row=row_idx, column=col_tr).value = tr or ''
            etiquetadas += 1

    print(f"Procesadas: {total}, etiquetadas: {etiquetadas}, COO->LBR: {coo_norm}, J31 vistas: {j31_seen}")
    print(f"Guardando {src.name}...")
    wb.save(src)
    wb.close()
    print(f"Listo en {time.time()-t0:.1f}s")

if __name__ == '__main__':
    main()
